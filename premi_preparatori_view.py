"""Interfaccia per il calcolo dei premi preparatori."""
import datetime
from contextlib import closing
from decimal import Decimal, ROUND_HALF_UP
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import mysql.connector

from config import COLORS, FONTS, MYSQL_CONFIG
from database import (
    fetch_fasce_premi,
    fetch_premi_preparatori,
    get_malus_bonus,
    save_premi_preparatori,
)
from ui_components import create_button


MONTH_CHOICES: List[Tuple[str, int]] = [
    ("Gennaio", 1),
    ("Febbraio", 2),
    ("Marzo", 3),
    ("Aprile", 4),
    ("Maggio", 5),
    ("Giugno", 6),
    ("Luglio", 7),
    ("Agosto", 8),
    ("Settembre", 9),
    ("Ottobre", 10),
    ("Novembre", 11),
    ("Dicembre", 12),
]


class PremiPreparatoriView(tk.Frame):
    """Interfaccia per il calcolo premi preparatori."""

    def __init__(self, parent: tk.Widget):
        super().__init__(parent, bg=COLORS["background"])
        self.pack(fill="both", expand=True)

        today = datetime.date.today()
        self.anno_var = tk.StringVar(value=str(today.year))
        self.mese_var = tk.StringVar(value=MONTH_CHOICES[today.month - 1][0])
        self.codice_var = tk.StringVar()
        self._current_premi: List[Dict[str, Any]] = []
        self._sort_reverse: Dict[str, bool] = {}

        self._build_ui()
        self._carica_premi()

    def _build_ui(self) -> None:
        """Costruisce l'interfaccia utente."""
        header = tk.Frame(self, bg=COLORS["background"])
        header.pack(fill="x", padx=16, pady=(16, 0))

        tk.Label(
            header,
            text="Calcolo Premi Preparatori",
            font=FONTS["title"],
            bg=COLORS["background"],
            fg=COLORS["primary"],
        ).pack(side="left")

        filter_frame = tk.Frame(self, bg=COLORS["background"], bd=2, relief="groove")
        filter_frame.pack(fill="x", padx=16, pady=12, ipady=12)

        tk.Label(
            filter_frame,
            text="Anno:",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).grid(row=0, column=0, sticky="w", padx=(12, 6), pady=10)

        current_year = datetime.date.today().year
        anni = [str(y) for y in range(current_year - 5, current_year + 2)]
        anno_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.anno_var,
            values=anni,
            width=12,
            state="readonly",
            font=FONTS["input"],
        )
        anno_combo.grid(row=0, column=1, sticky="ew", padx=(0, 16), pady=10)
        anno_combo.bind("<<ComboboxSelected>>", lambda event: self._carica_premi())

        tk.Label(
            filter_frame,
            text="Mese:",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).grid(row=0, column=2, sticky="w", padx=(0, 6), pady=10)

        mesi = [label for label, _ in MONTH_CHOICES]
        mese_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.mese_var,
            values=mesi,
            width=14,
            state="readonly",
            font=FONTS["input"],
        )
        mese_combo.grid(row=0, column=3, sticky="ew", padx=(0, 16), pady=10)
        mese_combo.bind("<<ComboboxSelected>>", lambda event: self._carica_premi())

        tk.Label(
            filter_frame,
            text="\U0001f50d Ricerca:",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).grid(row=0, column=4, sticky="w", padx=(0, 6), pady=10)

        search_entry = ttk.Entry(
            filter_frame,
            textvariable=self.codice_var,
            width=24,
            font=FONTS["input"],
        )
        search_entry.grid(row=0, column=5, sticky="ew", padx=(0, 16), pady=10)
        search_entry.bind("<Return>", lambda event: self._carica_premi())

        create_button(
            filter_frame,
            text="Aggiorna",
            command=self._carica_premi,
            variant="secondary",
            width=14,
        ).grid(row=0, column=6, padx=(12, 6), pady=10)

        create_button(
            filter_frame,
            text="Calcola Premi",
            command=self._genera_premi,
            variant="primary",
            width=16,
        ).grid(row=0, column=7, padx=(6, 12), pady=10)

        create_button(
            filter_frame,
            text="Export Excel",
            command=self._export_premi_excel,
            variant="secondary",
            width=14,
        ).grid(row=0, column=8, padx=(6, 12), pady=10)

        for col in [1, 3, 5]:
            filter_frame.grid_columnconfigure(col, weight=1)

        table_frame = tk.Frame(
            self,
            bg=COLORS["white"],
            highlightbackground=COLORS["border"],
            highlightthickness=1,
        )
        table_frame.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        columns = (
            "codice",
            "nome",
            "tot_colli",
            "ore",
            "colli_ora",
            "fascia",
            "premio_base",
            "penalita",
            "premio_kpi",
            "premio_totale",
        )

        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
        )

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        headers = {
            "codice": "Codice",
            "nome": "Nome",
            "tot_colli": "Tot. Colli",
            "ore": "Ore",
            "colli_ora": "Colli/h",
            "fascia": "Fascia",
            "premio_base": "Premio Base (EUR)",
            "penalita": "Penalità Totale (EUR)",
            "premio_kpi": "Premio KPI (EUR)",
            "premio_totale": "Premio Totale (EUR)",
        }

        self._tree_headers = headers

        for col, title in headers.items():
            self.tree.heading(col, text=title, command=lambda c=col: self._sort_by_column(c))

        self.tree.column("codice", width=90, anchor="center")
        self.tree.column("nome", width=180, anchor="w")
        self.tree.column("tot_colli", width=120, anchor="center")
        self.tree.column("ore", width=80, anchor="center")
        self.tree.column("colli_ora", width=100, anchor="center")
        self.tree.column("fascia", width=100, anchor="center")
        self.tree.column("premio_base", width=120, anchor="e")
        self.tree.column("penalita", width=110, anchor="e")
        self.tree.column("premio_kpi", width=120, anchor="e")
        self.tree.column("premio_totale", width=130, anchor="e")

        footer_frame = tk.Frame(
            self,
            bg=COLORS["primary"],
        )
        footer_frame.pack(fill="x", padx=16, pady=(4, 0))

        self.stats_label = tk.Label(
            footer_frame,
            text="",
            font=("Segoe UI", 12, "bold"),
            bg=COLORS["primary"],
            fg="white",
        )
        self.stats_label.pack(side="left", padx=12, pady=10)

    # --------------------------------------------------------- ORDINAMENTO ---
    def _sort_by_column(self, col: str) -> None:
        """Ordina la tabella in base alla colonna selezionata."""
        reverse = self._sort_reverse.get(col, False)
        items = list(self.tree.get_children(""))

        def sort_key(item_id: str):
            value = self.tree.set(item_id, col)
            return self._coerce_sort_value(value)

        items.sort(key=sort_key, reverse=reverse)

        for index, item_id in enumerate(items):
            self.tree.move(item_id, "", index)

        self._sort_reverse[col] = not reverse

    @staticmethod
    def _coerce_sort_value(value):
        if value is None:
            return (1, "")
        text = str(value).strip()
        if not text:
            return (1, "")
        normalized = text.replace(" ", "")
        if re.match(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$", normalized):
            normalized = normalized.replace(".", "").replace(",", ".")
        elif re.match(r"^-?\d+([\.,]\d+)?$", normalized):
            normalized = normalized.replace(",", ".")
        try:
            return (0, float(normalized))
        except ValueError:
            return (0, text.lower())

    def _export_premi_excel(self) -> None:
        """Esporta l'elenco premi corrente in Excel."""
        rows = self._collect_tree_rows()
        if not rows:
            messagebox.showinfo(
                "Nessun dato",
                "Non ci sono premi da esportare. Usa 'Aggiorna' o 'Calcola Premi'.",
                parent=self,
            )
            return

        anno = self.anno_var.get().strip() or "XXXX"
        mese = (self.mese_var.get().strip() or "Mese").replace(" ", "_")
        default_name = f"Premi_Preparatori_{anno}_{mese}.xlsx"

        file_path = filedialog.asksaveasfilename(
            title="Esporta premi in Excel",
            defaultextension=".xlsx",
            filetypes=[("File Excel", "*.xlsx"), ("Tutti i file", "*.*")],
            initialdir=str(Path.home()),
            initialfile=default_name,
        )

        if not file_path:
            return

        try:
            import pandas as pd
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment

            # Converte i valori stringa in numeri dove possibile
            numeric_cols = {
                "Tot. Colli", "Ore", "Colli/h",
                "Premio Base (EUR)", "Penalità Totale (EUR)",
                "Premio KPI (EUR)", "Premio Totale (EUR)",
            }

            def parse_numeric(val):
                if val is None or val == "":
                    return 0.0
                try:
                    # Rimuove eventuali separatori migliaia e converte
                    cleaned = str(val).replace(" ", "").replace(",", ".")
                    return float(cleaned)
                except (ValueError, TypeError):
                    return val

            processed_rows = []
            for row in rows:
                new_row = {}
                for key, val in row.items():
                    if key in numeric_cols:
                        new_row[key] = parse_numeric(val)
                    else:
                        new_row[key] = val
                processed_rows.append(new_row)

            df = pd.DataFrame(processed_rows)

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                sheet_name = "Premi Preparatori"
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                sheet = writer.sheets[sheet_name]

                # Formati numerici: "General" mostra il numero senza decimali inutili
                # Per i valori monetari/decimali usiamo formato condizionale
                format_int = "#,##0"           # Interi con separatore migliaia
                # Formato condizionale: se intero mostra senza decimali, altrimenti max 2
                format_auto = "0.00;-0.00;0"

                col_formats = {
                    "Tot. Colli": format_int,
                    "Ore": format_auto,
                    "Colli/h": format_auto,
                    "Premio Base (EUR)": format_auto,
                    "Penalità Totale (EUR)": format_auto,
                    "Premio KPI (EUR)": format_auto,
                    "Premio Totale (EUR)": format_auto,
                }

                right_align = Alignment(horizontal="right")

                # Auto-dimensiona le colonne e applica formati
                for col_idx, column_name in enumerate(df.columns, start=1):
                    column_letter = get_column_letter(col_idx)
                    # Larghezza = max tra header e valori + margine
                    max_length = len(str(column_name))
                    for cell in sheet[column_letter][1:]:
                        try:
                            cell_len = len(str(cell.value)) if cell.value else 0
                            if cell_len > max_length:
                                max_length = cell_len
                        except Exception:
                            pass
                    sheet.column_dimensions[column_letter].width = max_length + 3

                    # Applica formato numerico e allineamento
                    if column_name in col_formats:
                        for row_idx in range(2, len(df) + 2):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            val = cell.value
                            # Se è un intero, usa formato intero; altrimenti decimale
                            if column_name != "Tot. Colli":
                                try:
                                    if val is not None and float(val) == int(val):
                                        cell.number_format = format_int
                                    else:
                                        cell.number_format = format_auto
                                except (ValueError, TypeError):
                                    cell.number_format = format_auto
                            else:
                                cell.number_format = format_int
                            cell.alignment = right_align

                sheet.freeze_panes = "A2"

            open_folder = messagebox.askyesno(
                "Export completato",
                f"File salvato in:\n{file_path}\n\nAprire la cartella di destinazione?",
                parent=self,
            )
            if open_folder:
                try:
                    os.startfile(Path(file_path).parent)
                except OSError:
                    messagebox.showinfo(
                        "Cartella",
                        f"Apri manualmente la cartella:\n{Path(file_path).parent}",
                        parent=self,
                    )
        except Exception as exc:
            messagebox.showerror(
                "Errore",
                f"Esportazione non riuscita:\n{exc}",
                parent=self,
            )

    def _collect_tree_rows(self) -> List[Dict[str, Any]]:
        """Restituisce i dati mostrati nella tabella con header leggibili."""
        columns = list(self.tree["columns"])
        headers = [self._tree_headers.get(col, col) for col in columns]
        data: List[Dict[str, Any]] = []

        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            row = {
                headers[idx]: values[idx] if idx < len(values) else ""
                for idx in range(len(columns))
            }
            data.append(row)

        return data

    def _carica_premi(self) -> None:
        """Carica i premi salvati dal database."""
        anno_str = self.anno_var.get().strip()
        mese_label = self.mese_var.get().strip()
        search_text = self.codice_var.get().strip() or None

        if not anno_str or not mese_label:
            return

        try:
            anno = int(anno_str)
            mese = next(m for label, m in MONTH_CHOICES if label == mese_label)
        except (ValueError, StopIteration):
            return

        try:
            premi = fetch_premi_preparatori(anno, mese, search=search_text)
            self._current_premi = premi.copy()

            for item in self.tree.get_children():
                self.tree.delete(item)

            totale_premi = Decimal("0")
            totale_penalita = Decimal("0")
            bonus_applicato = False

            for premio in premi:
                colli = Decimal(str(premio.get("totale_colli") or 0))
                ore = Decimal(str(premio.get("ore_lavorate") or 0))
                colli_ora = Decimal(str(premio.get("colli_ora") or 0))
                premio_base = Decimal(str(premio.get("premio_base") or 0))
                penalita = Decimal(str(premio.get("penalita_totale") or 0))
                premio_kpi = Decimal(str(premio.get("premio_kpi") or 0))
                premio_totale = Decimal(str(premio.get("premio_totale") or 0))

                netto = premio_base - penalita
                if netto < Decimal("0"):
                    netto = Decimal("0")

                if premio.get("bonus_applicato") or premio_kpi > 0:
                    bonus_applicato = True

                self.tree.insert(
                    "",
                    "end",
                    values=(
                        premio.get("codice_preparatore"),
                        premio.get("nome_preparatore"),
                        f"{colli:.0f}",
                        f"{ore:.2f}",
                        f"{colli_ora:.2f}",
                        premio.get("fascia_raggiunta") or "N/A",
                        f"{premio_base:.2f}",
                        f"{penalita:.2f}",
                        f"{premio_kpi:.2f}",
                        f"{premio_totale:.2f}",
                    ),
                )

                totale_premi += premio_totale
                totale_penalita += penalita

            if not premi:
                self.stats_label.config(
                    text=f"Nessun premio calcolato per {mese_label} {anno}. Usa 'Calcola Premi' per generarli."
                )
            else:
                bonus_text = "SI" if bonus_applicato else "NO"
                self.stats_label.config(
                    text=(
                        f"Preparatori: {len(premi)} | Totale Premi: EUR {totale_premi:,.2f}"
                        f" | Penalita: EUR {totale_penalita:,.2f} | Bonus KPI: {bonus_text}"
                    )
                )

        except Exception as exc:
            messagebox.showerror(
                "Errore",
                f"Errore nel caricamento premi:\n{exc}",
                parent=self,
            )

    def _genera_premi(self) -> None:
        """Calcola e salva i premi per il periodo selezionato."""
        anno_str = self.anno_var.get().strip()
        mese_label = self.mese_var.get().strip()

        if not anno_str or not mese_label:
            messagebox.showwarning(
                "Dati mancanti",
                "Seleziona anno e mese.",
                parent=self,
            )
            return

        try:
            anno = int(anno_str)
            mese = next(m for label, m in MONTH_CHOICES if label == mese_label)
        except (ValueError, StopIteration):
            messagebox.showerror("Errore", "Anno o mese non validi.", parent=self)
            return

        premi_esistenti = fetch_premi_preparatori(anno, mese)
        if premi_esistenti:
            risposta = messagebox.askyesno(
                "Conferma",
                (
                    f"Esistono gia {len(premi_esistenti)} premi calcolati per {mese_label} {anno}.\n\n"
                    "Vuoi ricalcolarli? I dati precedenti verranno sovrascritti."
                ),
                parent=self,
            )
            if not risposta:
                return

        try:
            fasce = self._load_fasce_premio()
            if not fasce:
                messagebox.showwarning(
                    "Fasce mancanti",
                    "Non sono definite fasce premio per PICKING.",
                    parent=self,
                )
                return

            bonus_perc = self._load_bonus_malus(anno, mese)

            risultati = self._calcola_premi_preparatori(
                anno=anno,
                mese=mese,
                codice_filtro=None,
                fasce=fasce,
                bonus_perc=bonus_perc,
            )

            if not risultati:
                messagebox.showinfo(
                    "Nessun dato",
                    f"Nessun dato di produzione trovato per {mese_label} {anno}.",
                    parent=self,
                )
                return

            for ris in risultati:
                ris["bonus_applicato"] = bool(bonus_perc and ris.get("premio_kpi", 0) > 0)

            save_premi_preparatori(anno, mese, risultati)

            messagebox.showinfo(
                "Successo",
                f"Calcolati e salvati {len(risultati)} premi per {mese_label} {anno}.",
                parent=self,
            )

            self._carica_premi()
        except Exception as exc:
            messagebox.showerror(
                "Errore",
                f"Errore nel calcolo premi:\n{exc}",
                parent=self,
            )

    def _load_fasce_premio(self) -> List[Dict]:
        """Recupera le fasce premio per i preparatori."""
        fasce = fetch_fasce_premi("PICKING")
        return sorted(fasce, key=lambda fascia: fascia.get("valore_riferimento", 0))

    def _load_bonus_malus(self, anno: int, mese: int) -> Optional[Decimal]:
        """Recupera la percentuale di bonus applicabile per il mese."""
        record = get_malus_bonus(anno, mese)
        if not record:
            return None

        attivita_bonus = record.get("attivita_bonus", "")
        if not attivita_bonus or "PICKING" not in attivita_bonus.upper():
            return None

        try:
            rotture = Decimal(str(record.get("importo_rotture", 0) or 0))
            differenze = Decimal(str(record.get("importo_differenze", 0) or 0))
            soglia_rot = Decimal(str(record.get("soglia_rotture", 0) or 0))
            soglia_diff = Decimal(str(record.get("soglia_differenze", 0) or 0))

            # Il bonus scatta solo se ENTRAMBE le soglie sono rispettate
            # (verifica separata, non aggregata)
            if rotture <= soglia_rot and differenze <= soglia_diff:
                return Decimal("0.15")
        except Exception:
            return None

        return None

    def _calcola_premi_preparatori(
        self,
        anno: int,
        mese: int,
        codice_filtro: Optional[str],
        fasce: List[Dict],
        bonus_perc: Optional[Decimal],
    ) -> List[Dict]:
        """Calcola i premi per i preparatori."""
        import mysql.connector
        from contextlib import closing
        from config import MYSQL_CONFIG
        from typing import Any, cast

        cent = Decimal("0.01")

        query = """
            SELECT
                dp.codice_preparatore,
                dp.nome_preparatore,
                SUM(dp.totale_colli) AS totale_colli,
                SUM(dp.ore_tim) AS ore_tim,
                SUM(dp.ore_gestionale) AS ore_gestionale,
                SUM(dp.penalita_totale) AS penalita_totale
            FROM dati_produzione dp
            WHERE dp.tipo_attivita = 'PICKING'
                AND YEAR(dp.data) = %s
                AND MONTH(dp.data) = %s
                AND NOT EXISTS (
                    SELECT 1 FROM anomalie a
                    WHERE a.tipo_anomalia IN ('PRODUZIONE_SENZA_ORE', 'ORE_SENZA_PRODUZIONE')
                        AND a.data_rilevamento = dp.data
                        AND a.codice_preparatore = dp.codice_preparatore
                        AND a.tipo_attivita = dp.tipo_attivita
                        AND COALESCE(a.stato, 'APERTA') NOT IN ('RISOLTA', 'VERIFICATA')
                )
        """
        params: List[Any] = [anno, mese]

        if codice_filtro:
            query += " AND dp.codice_preparatore = %s"
            params.append(codice_filtro)

        query += " GROUP BY dp.codice_preparatore, dp.nome_preparatore"

        risultati: List[Dict[str, Any]] = []

        with closing(mysql.connector.connect(**MYSQL_CONFIG)) as conn:
            with closing(conn.cursor(dictionary=True)) as cur:
                cur.execute(query, params)
                rows = cast(List[Dict[str, Any]], cur.fetchall())

                for row in rows:
                    codice = str(row.get("codice_preparatore") or "")
                    nome = str(row.get("nome_preparatore") or "")
                    totale_colli = Decimal(str(row.get("totale_colli") or 0))
                    # ⚠️ IMPORTANTE: ore_tim e ore_gestionale sono in MINUTI nel DB
                    minuti_tim = Decimal(str(row.get("ore_tim") or 0))
                    minuti_gest = Decimal(str(row.get("ore_gestionale") or 0))
                    penalita = Decimal(str(row.get("penalita_totale") or 0))

                    # Converti minuti → ore
                    ore_tim = minuti_tim / Decimal("60")
                    ore_gest = minuti_gest / Decimal("60")
                    
                    # I premi preparatori utilizzano le ore_tim come riferimento orario
                    ore_effettive = ore_tim
                    if ore_effettive <= 0:
                        continue

                    colli_ora = (totale_colli / ore_effettive).quantize(Decimal("0.01"))

                    fascia_label = "N/A"
                    premio_unitario = Decimal("0")
                    for fascia in fasce:
                        soglia = Decimal(str(fascia.get("valore_riferimento", 0)))
                        if colli_ora >= soglia:
                            premio_unitario = Decimal(str(fascia.get("valore_premio", 0)))
                            unita = fascia.get("unita_riferimento") or "Colli/h"
                            fascia_label = f"{soglia} {unita}"

                    premio_base = (premio_unitario * totale_colli).quantize(cent)

                    premio_netto = premio_base - penalita
                    if premio_netto < Decimal("0"):
                        premio_netto = Decimal("0")

                    premio_kpi = Decimal("0")
                    if bonus_perc and premio_netto > 0:
                        premio_kpi = (premio_netto * bonus_perc).quantize(cent)

                    premio_totale = (premio_netto + premio_kpi).quantize(cent)

                    colli_int = int(
                        totale_colli.to_integral_value(rounding=ROUND_HALF_UP)
                    )

                    risultati.append(
                        {
                            "codice": codice,
                            "nome": nome,
                            "tot_colli": colli_int,
                            "ore": float(ore_effettive),
                            "colli_ora": float(colli_ora),
                            "fascia": fascia_label,
                            "premio_base": float(premio_base),
                            "penalita": float(penalita.quantize(cent)),
                            "premio_kpi": float(premio_kpi),
                            "premio_totale": float(premio_totale),
                            "note": None,
                        }
                    )

        risultati.sort(key=lambda item: item["premio_totale"], reverse=True)
        return risultati