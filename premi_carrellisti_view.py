"""
Interfaccia per il calcolo dei premi carrellisti.
"""
import datetime
from decimal import Decimal
import os
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from config import COLORS, FONTS
from database import (
    fetch_fasce_premi,
    get_malus_bonus,
    fetch_pesi_movimenti,
    fetch_premi_carrellisti,
    save_premi_carrellisti,
    delete_premi_carrellisti,
)
from ui_components import create_button


MONTH_CHOICES: List[Tuple[str, int]] = [
    ("Gennaio", 1),
    ("Febbraio", 2),
    ("Marzo", 3),
    ("Aprile", 4),
    ("Maggio", 5),
    ("Giugno", 6),
    ("Luglio", 7),
    ("Agosto", 8),
    ("Settembre", 9),
    ("Ottobre", 10),
    ("Novembre", 11),
    ("Dicembre", 12),
]


class PremiCarrellistiView(tk.Frame):
    """Interfaccia per il calcolo premi carrellisti."""

    def __init__(self, parent: tk.Widget):
        super().__init__(parent, bg=COLORS["background"])
        self.pack(fill="both", expand=True)

        today = datetime.date.today()
        self.anno_var = tk.StringVar(value=str(today.year))
        self.mese_var = tk.StringVar(value=MONTH_CHOICES[today.month - 1][0])
        self.codice_var = tk.StringVar()
        self._current_premi: List[Dict[str, Any]] = []
        self._sort_reverse: Dict[str, bool] = {}

        self._build_ui()
        # Carica automaticamente i premi del mese corrente
        self._carica_premi()

    def _build_ui(self):
        """Costruisce l'interfaccia utente."""
        # Header
        header = tk.Frame(self, bg=COLORS["background"])
        header.pack(fill="x", padx=16, pady=(16, 0))

        tk.Label(
            header,
            text="🏆 Calcolo Premi Carrellisti",
            font=FONTS["title"],
            bg=COLORS["background"],
            fg=COLORS["primary"],
        ).pack(side="left")

        # Filtri
        filter_frame = tk.Frame(self, bg=COLORS["background"], bd=2, relief="groove")
        filter_frame.pack(fill="x", padx=16, pady=12, ipady=12)

        # Anno
        tk.Label(
            filter_frame,
            text="Anno:",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).grid(row=0, column=0, sticky="w", padx=(12, 6), pady=10)

        current_year = datetime.date.today().year
        anni = [str(y) for y in range(current_year - 5, current_year + 2)]
        anno_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.anno_var,
            values=anni,
            width=12,
            state="readonly",
            font=FONTS["input"],
        )
        anno_combo.grid(row=0, column=1, sticky="ew", padx=(0, 16), pady=10)
        anno_combo.bind("<<ComboboxSelected>>", lambda e: self._carica_premi())

        # Mese
        tk.Label(
            filter_frame,
            text="Mese:",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).grid(row=0, column=2, sticky="w", padx=(0, 6), pady=10)

        mesi = [label for label, _ in MONTH_CHOICES]
        mese_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.mese_var,
            values=mesi,
            width=14,
            state="readonly",
            font=FONTS["input"],
        )
        mese_combo.grid(row=0, column=3, sticky="ew", padx=(0, 16), pady=10)
        mese_combo.bind("<<ComboboxSelected>>", lambda e: self._carica_premi())

        # Codice (opzionale)
        tk.Label(
            filter_frame,
            text="Codice (opzionale):",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).grid(row=0, column=4, sticky="w", padx=(0, 6), pady=10)

        codice_entry = ttk.Entry(
            filter_frame,
            textvariable=self.codice_var,
            width=16,
            font=FONTS["input"],
        )
        codice_entry.grid(row=0, column=5, sticky="ew", padx=(0, 16), pady=10)
        codice_entry.bind("<Return>", lambda e: self._carica_premi())

        # Pulsanti
        create_button(
            filter_frame,
            text="🔄 Aggiorna",
            command=self._carica_premi,
            variant="secondary",
            width=14,
        ).grid(row=0, column=6, padx=(12, 6), pady=10)

        create_button(
            filter_frame,
            text="🎯 Calcola Premi",
            command=self._genera_premi,
            variant="primary",
            width=16,
        ).grid(row=0, column=7, padx=(6, 12), pady=10)

        create_button(
            filter_frame,
            text="📤 Export Excel",
            command=self._export_premi_excel,
            variant="secondary",
            width=16,
        ).grid(row=0, column=8, padx=(6, 12), pady=10)

        # Configura espansione colonne
        for col in [1, 3, 5]:
            filter_frame.grid_columnconfigure(col, weight=1)

        # Tabella risultati
        table_frame = tk.Frame(
            self,
            bg=COLORS["white"],
            highlightbackground=COLORS["border"],
            highlightthickness=1,
        )
        table_frame.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        columns = (
            "codice",
            "nome",
            "tot_movimenti",
            "ore",
            "mov_ora",
            "fascia",
            "premio_base",
            "premio_kpi",
            "premio_totale",
        )

        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
        )

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Headers
        headers = {
            "codice": "Codice",
            "nome": "Nome",
            "tot_movimenti": "Tot. Movimenti",
            "ore": "Ore",
            "mov_ora": "Mov/h",
            "fascia": "Fascia",
            "premio_base": "Premio Base (€)",
            "premio_kpi": "Premio KPI (€)",
            "premio_totale": "Premio Totale (€)",
        }

        self._tree_headers = headers

        for col, title in headers.items():
            self.tree.heading(col, text=title, command=lambda c=col: self._sort_by_column(c))

        self.tree.column("codice", width=90, anchor="center")
        self.tree.column("nome", width=180, anchor="w")
        self.tree.column("tot_movimenti", width=120, anchor="center")
        self.tree.column("ore", width=80, anchor="center")
        self.tree.column("mov_ora", width=100, anchor="center")
        self.tree.column("fascia", width=100, anchor="center")
        self.tree.column("premio_base", width=120, anchor="e")
        self.tree.column("premio_kpi", width=120, anchor="e")
        self.tree.column("premio_totale", width=130, anchor="e")

        # Footer con statistiche
        footer_frame = tk.Frame(
            self,
            bg=COLORS["white"],
            highlightbackground=COLORS["border"],
            highlightthickness=1,
        )
        footer_frame.pack(fill="x", padx=16)

        self.stats_label = tk.Label(
            footer_frame,
            text="",
            font=FONTS.get("subtitle", ("Segoe UI", 10)),
            bg=COLORS["white"],
            fg=COLORS["text_light"],
        )
        self.stats_label.pack(side="left", padx=12, pady=10)

    # --------------------------------------------------------- ORDINAMENTO ---
    def _sort_by_column(self, col: str) -> None:
        """Ordina la tabella in base alla colonna selezionata."""
        reverse = self._sort_reverse.get(col, False)
        items = list(self.tree.get_children(""))

        def sort_key(item_id: str):
            value = self.tree.set(item_id, col)
            return self._coerce_sort_value(value)

        items.sort(key=sort_key, reverse=reverse)

        for index, item_id in enumerate(items):
            self.tree.move(item_id, "", index)

        self._sort_reverse[col] = not reverse

    @staticmethod
    def _coerce_sort_value(value):
        if value is None:
            return (1, "")
        text = str(value).strip()
        if not text:
            return (1, "")
        normalized = text.replace(" ", "")
        if re.match(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$", normalized):
            normalized = normalized.replace(".", "").replace(",", ".")
        elif re.match(r"^-?\d+([\.,]\d+)?$", normalized):
            normalized = normalized.replace(",", ".")
        try:
            return (0, float(normalized))
        except ValueError:
            return (0, text.lower())

    def _export_premi_excel(self) -> None:
        """Esporta i premi mostrati nella tabella in formato Excel."""
        premi_src = self._current_premi.copy() if self._current_premi else []
        if not premi_src and not self.tree.get_children():
            messagebox.showinfo(
                "Nessun dato",
                "Non ci sono premi da esportare. Usa 'Aggiorna' o 'Calcola Premi'.",
                parent=self,
            )
            return

        anno = self.anno_var.get().strip() or "XXXX"
        mese = (self.mese_var.get().strip() or "Mese").replace(" ", "_")
        default_name = f"Premi_Carrellisti_{anno}_{mese}.xlsx"

        file_path = filedialog.asksaveasfilename(
            title="Esporta premi in Excel",
            defaultextension=".xlsx",
            filetypes=[("File Excel", "*.xlsx"), ("Tutti i file", "*.*")],
            initialdir=str(Path.home()),
            initialfile=default_name,
        )

        if not file_path:
            return

        try:
            import pandas as pd

            cent_cols = {
                "Ore",
                "Mov/h",
                "Premio Base (€)",
                "Premio KPI (€)",
                "Premio Totale (€)",
            }
            int_cols = {"Tot. Movimenti"}

            if premi_src:
                export_rows: List[Dict[str, Any]] = []
                for premio in premi_src:
                    export_rows.append(
                        {
                            "Codice": (premio.get("codice_preparatore") or ""),
                            "Nome": (premio.get("nome_preparatore") or ""),
                            "Tot. Movimenti": int(float(premio.get("totale_movimenti") or 0)),
                            "Ore": float(premio.get("ore_lavorate") or 0),
                            "Mov/h": float(premio.get("movimenti_ora") or 0),
                            "Fascia": (premio.get("fascia_raggiunta") or "N/A"),
                            "Premio Base (€)": float(premio.get("premio_base") or 0),
                            "Premio KPI (€)": float(premio.get("premio_kpi") or 0),
                            "Premio Totale (€)": float(premio.get("premio_totale") or 0),
                        }
                    )
                df = pd.DataFrame(export_rows)
            else:
                rows = self._collect_tree_rows()
                df = pd.DataFrame(rows)

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Premi")
                ws = writer.sheets["Premi"]

                headers = [cell.value for cell in ws[1]]
                col_index: Dict[str, int] = {
                    str(h): i + 1 for i, h in enumerate(headers) if h is not None
                }

                for col_name, idx in col_index.items():
                    if col_name in int_cols:
                        fmt = "0"
                    elif col_name in cent_cols:
                        fmt = "#,##0.00"
                    else:
                        continue

                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=idx)
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            cell.number_format = fmt

                from openpyxl.utils import get_column_letter

                for col in range(1, ws.max_column + 1):
                    letter = get_column_letter(col)
                    max_len = 0
                    for row_idx in range(1, ws.max_row + 1):
                        value = ws.cell(row=row_idx, column=col).value
                        if value is None:
                            continue
                        max_len = max(max_len, len(str(value)))
                    ws.column_dimensions[letter].width = min(max_len + 2, 60)
            open_folder = messagebox.askyesno(
                "Export completato",
                f"File salvato in:\n{file_path}\n\nAprire la cartella di destinazione?",
                parent=self,
            )
            if open_folder:
                try:
                    os.startfile(Path(file_path).parent)
                except OSError:
                    messagebox.showinfo(
                        "Cartella",
                        f"Apri manualmente la cartella:\n{Path(file_path).parent}",
                        parent=self,
                    )
        except Exception as exc:
            messagebox.showerror(
                "Errore",
                f"Esportazione non riuscita:\n{exc}",
                parent=self,
            )

    def _collect_tree_rows(self) -> List[Dict[str, str]]:
        """Raccoglie i dati visibili nella treeview."""
        columns = list(self.tree["columns"])
        headers = [self._tree_headers.get(col, col) for col in columns]
        data: List[Dict[str, str]] = []

        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            row = {
                headers[idx]: values[idx] if idx < len(values) else ""
                for idx in range(len(columns))
            }
            data.append(row)

        return data

    def _carica_premi(self):
        """Carica e visualizza i premi salvati nel database."""
        anno_str = self.anno_var.get().strip()
        mese_label = self.mese_var.get().strip()
        codice_filtro = self.codice_var.get().strip().upper() or None

        if not anno_str or not mese_label:
            return

        try:
            anno = int(anno_str)
            mese = next(m for label, m in MONTH_CHOICES if label == mese_label)
        except (ValueError, StopIteration):
            return

        try:
            # Recupera premi dal database
            premi = fetch_premi_carrellisti(anno, mese, codice_filtro)
            self._current_premi = premi.copy()
            
            # Pulisci tabella
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Mostra risultati
            totale_premi = 0.0
            bonus_applicato = False

            for premio in premi:
                mov = float(premio.get("totale_movimenti") or 0)
                ore = float(premio.get("ore_lavorate") or 0)
                mov_ora = float(premio.get("movimenti_ora") or 0)
                fascia = str(premio.get("fascia_raggiunta") or "N/A")
                base = float(premio.get("premio_base") or 0)
                kpi = float(premio.get("premio_kpi") or 0)
                totale = float(premio.get("premio_totale") or 0)
                
                if premio.get("bonus_applicato"):
                    bonus_applicato = True

                self.tree.insert(
                    "",
                    "end",
                    values=(
                        premio.get("codice_preparatore"),
                        premio.get("nome_preparatore"),
                        f"{mov:.0f}",
                        f"{ore:.2f}",
                        f"{mov_ora:.2f}",
                        fascia,
                        f"{base:.2f}",
                        f"{kpi:.2f}",
                        f"{totale:.2f}",
                    ),
                )
                totale_premi += totale

            # Aggiorna statistiche
            bonus_text = f" | Bonus KPI 15%: {'✓' if bonus_applicato else '✗'}"
            if len(premi) == 0:
                self.stats_label.config(
                    text=f"Nessun premio calcolato per {mese_label} {anno}. Usa 'Calcola Premi' per generarli."
                )
            else:
                self.stats_label.config(
                    text=f"Carrellisti: {len(premi)} | Totale Premi: €{totale_premi:,.2f}{bonus_text}"
                )

        except Exception as exc:
            messagebox.showerror(
                "Errore",
                f"Errore nel caricamento premi:\n{exc}",
                parent=self,
            )

    def _genera_premi(self):
        """Genera il calcolo dei premi per il periodo selezionato e li salva nel database."""
        anno_str = self.anno_var.get().strip()
        mese_label = self.mese_var.get().strip()

        if not anno_str or not mese_label:
            messagebox.showwarning(
                "Dati mancanti",
                "Seleziona anno e mese.",
                parent=self,
            )
            return

        try:
            anno = int(anno_str)
            mese = next(m for label, m in MONTH_CHOICES if label == mese_label)
        except (ValueError, StopIteration):
            messagebox.showerror("Errore", "Anno o mese non validi.", parent=self)
            return

        # Conferma sovrascrittura se esistono già premi
        premi_esistenti = fetch_premi_carrellisti(anno, mese)
        if premi_esistenti:
            risposta = messagebox.askyesno(
                "Conferma",
                f"Esistono già {len(premi_esistenti)} premi calcolati per {mese_label} {anno}.\n\n"
                f"Vuoi ricalcolarli? I dati precedenti verranno sovrascritti.",
                parent=self,
            )
            if not risposta:
                return

        try:
            # 1. Carica pesi movimenti per CARRELLISTI
            pesi_map = self._load_pesi_movimenti()

            # 2. Carica fasce premio per CARRELLISTI
            fasce = self._load_fasce_premio()

            # 3. Carica bonus/malus del mese
            bonus_perc = self._load_bonus_malus(anno, mese)

            # 4. Calcola dati produzione
            risultati = self._calcola_premi_carrellisti(
                anno, mese, None, pesi_map, fasce, bonus_perc
            )

            if not risultati:
                messagebox.showinfo(
                    "Nessun dato",
                    f"Nessun dato di produzione trovato per {mese_label} {anno}.",
                    parent=self,
                )
                return

            # Aggiungi flag bonus_applicato
            for ris in risultati:
                ris["bonus_applicato"] = bool(bonus_perc and ris.get("premio_kpi", 0) > 0)

            # 5. Salva nel database
            save_premi_carrellisti(anno, mese, risultati)

            messagebox.showinfo(
                "Successo",
                f"Calcolati e salvati {len(risultati)} premi per {mese_label} {anno}.",
                parent=self,
            )

            # 6. Ricarica la visualizzazione
            self._carica_premi()

        except Exception as exc:
            messagebox.showerror(
                "Errore",
                f"Errore nel calcolo premi:\n{exc}",
                parent=self,
            )

    def _load_pesi_movimenti(self) -> Dict[str, Decimal]:
        """Carica i pesi per tipo movimento dei carrellisti."""
        pesi = fetch_pesi_movimenti("CARRELLISTI")
        pesi_map = {}
        for peso in pesi:
            tipo = str(peso.get("tipo", "")).upper()
            valore = peso.get("peso")
            if tipo and valore is not None:
                pesi_map[tipo] = Decimal(str(valore))
        return pesi_map

    def _load_fasce_premio(self) -> List[Dict]:
        """Carica le fasce premio per carrellisti."""
        fasce = fetch_fasce_premi("CARRELLISTI")
        # Ordina per valore_riferimento crescente
        return sorted(fasce, key=lambda f: f.get("valore_riferimento", 0))

    def _load_bonus_malus(self, anno: int, mese: int) -> Optional[Decimal]:
        """Carica la percentuale bonus/malus per il mese."""
        record = get_malus_bonus(anno, mese)
        if not record:
            return None

        # Verifica se CARRELLISTI è nelle attività bonus
        attivita_bonus = record.get("attivita_bonus", "")
        if not attivita_bonus or "CARRELLISTI" not in attivita_bonus.upper():
            return None

        # Calcola bonus totale
        try:
            rotture = Decimal(str(record.get("importo_rotture", 0) or 0))
            differenze = Decimal(str(record.get("importo_differenze", 0) or 0))
            soglia_rot = Decimal(str(record.get("soglia_rotture", 0) or 0))
            soglia_diff = Decimal(str(record.get("soglia_differenze", 0) or 0))

            # Il bonus scatta solo se ENTRAMBE le soglie sono rispettate
            # (verifica separata, non aggregata)
            if rotture <= soglia_rot and differenze <= soglia_diff:
                return Decimal("0.15")
            else:
                return None
        except Exception:
            return None

    def _calcola_premi_carrellisti(
        self,
        anno: int,
        mese: int,
        codice_filtro: Optional[str],
        pesi_map: Dict[str, Decimal],
        fasce: List[Dict],
        bonus_perc: Optional[Decimal],
    ) -> List[Dict]:
        """Calcola i premi per tutti i carrellisti."""
        import mysql.connector
        from contextlib import closing
        from config import MYSQL_CONFIG
        from typing import Any, cast

        cent = Decimal("0.01")

        # Query per recuperare dati produzione
        # ESCLUDE i giorni con anomalia PRODUZIONE_SENZA_ORE
        query = """
            SELECT 
                dp.codice_preparatore,
                dp.nome_preparatore,
                dp.tipo,
                SUM(dp.totale_colli) as colli,
                SUM(dp.ore_tim) as ore_totali
            FROM dati_produzione dp
            WHERE dp.tipo_attivita = 'CARRELLISTI'
                AND YEAR(dp.data) = %s
                AND MONTH(dp.data) = %s
                AND NOT EXISTS (
                    SELECT 1 FROM anomalie a
                    WHERE a.tipo_anomalia IN ('PRODUZIONE_SENZA_ORE', 'ORE_SENZA_PRODUZIONE')
                        AND a.data_rilevamento = dp.data
                        AND a.codice_preparatore = dp.codice_preparatore
                        AND a.tipo_attivita = dp.tipo_attivita
                        AND COALESCE(a.stato, 'APERTA') NOT IN ('RISOLTA', 'VERIFICATA')
                )
        """
        params: List[Any] = [anno, mese]

        if codice_filtro:
            query += " AND dp.codice_preparatore = %s"
            params.append(codice_filtro)

        query += " GROUP BY dp.codice_preparatore, dp.nome_preparatore, dp.tipo"

        risultati_utente = {}

        with closing(mysql.connector.connect(**MYSQL_CONFIG)) as conn:
            with closing(conn.cursor(dictionary=True)) as cur:
                cur.execute(query, params)
                rows = cast(List[Dict[str, Any]], cur.fetchall())

                for row in rows:
                    codice = str(row.get("codice_preparatore") or "")
                    nome = str(row.get("nome_preparatore") or "")
                    tipo = str(row.get("tipo") or "").upper()
                    colli = int(row.get("colli") or 0)
                    # ⚠️ IMPORTANTE: ore_totali è in MINUTI nel DB
                    minuti_totali = Decimal(str(row.get("ore_totali") or 0))
                    # Converti minuti → ore
                    ore = minuti_totali / Decimal("60")

                    # Applica peso
                    peso = pesi_map.get(tipo, Decimal("1.0"))
                    movimenti_pesati = Decimal(str(colli)) * peso

                    key = codice
                    if key not in risultati_utente:
                        risultati_utente[key] = {
                            "codice": codice,
                            "nome": nome,
                            "movimenti_totali": Decimal("0"),
                            "ore_totali": Decimal("0"),
                        }

                    risultati_utente[key]["movimenti_totali"] += movimenti_pesati
                    risultati_utente[key]["ore_totali"] += ore

        # Calcola premi per ogni utente
        risultati_finali = []
        for dati in risultati_utente.values():
            movimenti = dati["movimenti_totali"]
            ore = dati["ore_totali"]

            if ore == 0:
                continue

            mov_ora = movimenti / ore

            # Trova fascia premio
            fascia_premio = None
            premio_base = Decimal("0")
            premio_unitario = Decimal("0")
            fascia_label = "N/A"

            for fascia in fasce:
                valore_rif = Decimal(str(fascia.get("valore_riferimento", 0)))
                if mov_ora >= valore_rif:
                    fascia_premio = fascia
                    premio_unitario = Decimal(str(fascia.get("valore_premio", 0)))
                    fascia_label = f"{valore_rif} Mov/h"

            if fascia_premio and premio_unitario > 0:
                premio_base = (premio_unitario * movimenti).quantize(cent)
            else:
                premio_base = Decimal("0")

            # Calcola premio KPI (15% se c'è bonus)
            premio_kpi = Decimal("0")
            if bonus_perc and premio_base > 0:
                premio_kpi = (premio_base * bonus_perc).quantize(cent)

            premio_totale = premio_base + premio_kpi
            premio_totale = premio_totale.quantize(cent)

            risultati_finali.append({
                "codice": dati["codice"],
                "nome": dati["nome"],
                "tot_movimenti": float(movimenti),
                "ore": float(ore),
                "mov_ora": float(mov_ora),
                "fascia": fascia_label,
                "premio_base": float(premio_base),
                "premio_kpi": float(premio_kpi),
                "premio_totale": float(premio_totale),
            })

        # Ordina per premio totale decrescente
        risultati_finali.sort(key=lambda x: x["premio_totale"], reverse=True)

        return risultati_finali
