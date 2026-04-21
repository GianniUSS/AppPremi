"""Interfaccia per il calcolo dei premi Ricevitori."""
import calendar
import os
import datetime
from contextlib import closing
from decimal import Decimal, ROUND_HALF_UP
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import mysql.connector

from config import COLORS, FONTS, MYSQL_CONFIG, MYSQL_CONFIG_MAIN
from database import (
    fetch_fasce_premi,
    fetch_premi_ricevitori,
    get_malus_bonus,
    save_premi_ricevitori,
)
from ui_components import create_button

MONTH_CHOICES: List[Tuple[str, int]] = [
    ("Gennaio", 1),
    ("Febbraio", 2),
    ("Marzo", 3),
    ("Aprile", 4),
    ("Maggio", 5),
    ("Giugno", 6),
    ("Luglio", 7),
    ("Agosto", 8),
    ("Settembre", 9),
    ("Ottobre", 10),
    ("Novembre", 11),
    ("Dicembre", 12),
]

# Varianti di descrizioni possibili nel gestionale per il reparto ricevimento
LOCAL_TIPI_UFF = ("UFF. RICEVIMENTO", "UFF.RICEVIMENTO", "UFF RICEVIMENTO")
TIM_TIPI_MAG = ("MAG. RICEVIMENTO", "MAG.RICEVIMENTO", "MAG RICEVIMENTO")
TIM_TIPI_UFF = ("UFF. RICEVIMENTO", "UFF.RICEVIMENTO", "UFF RICEVIMENTO")
DEBUG_PREPARATORI_ORE = {"NOTTE GIUSEPPE"}


class PremiRicevitoriView(tk.Frame):
    """Modulo UI per generare e visualizzare i premi del reparto ricevimento."""

    _saved_state: Dict[str, Any] = {}

    def __init__(self, parent: tk.Widget):
        super().__init__(parent, bg=COLORS["background"])
        self.pack(fill="both", expand=True)

        import app_state
        s = PremiRicevitoriView._saved_state
        self.anno_var = tk.StringVar(value=str(app_state.get_anno()))
        self.mese_var = tk.StringVar(value=app_state.get_mese_label())
        self.codice_var = tk.StringVar(value=s.get("search", ""))
        self.ore_giorno_var = tk.StringVar(value="8")
        self.total_ore_mag_var = tk.StringVar(value="0.00")
        self.total_ore_uff_var = tk.StringVar(value="0.00")
        self.total_pallet_var = tk.StringVar(value="0")
        self.media_pallet_ora_var = tk.StringVar(value="0.00")
        self._current_premi: List[Dict[str, Any]] = []
        self._sort_reverse: Dict[str, bool] = {}

        self._build_ui()
        self._carica_premi()

    # ------------------------------------------------------------------ UI ---
    def _build_ui(self) -> None:
        header = tk.Frame(self, bg=COLORS["background"])
        header.pack(fill="x", padx=16, pady=(16, 0))

        tk.Label(
            header,
            text="🏆 Calcolo Premi Ricevitori",
            font=FONTS["title"],
            bg=COLORS["background"],
            fg=COLORS["primary"],
        ).pack(side="left")

        filter_frame = tk.Frame(self, bg=COLORS["background"], bd=2, relief="groove")
        filter_frame.pack(fill="x", padx=16, pady=12, ipady=12)

        # Anno
        tk.Label(
            filter_frame,
            text="Anno:",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).grid(row=0, column=0, sticky="w", padx=(12, 6), pady=10)

        current_year = datetime.date.today().year
        anni = [str(y) for y in range(current_year - 5, current_year + 2)]
        anno_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.anno_var,
            values=anni,
            width=12,
            state="readonly",
            font=FONTS["input"],
        )
        anno_combo.grid(row=0, column=1, sticky="ew", padx=(0, 16), pady=10)
        anno_combo.bind("<<ComboboxSelected>>", lambda _: self._carica_premi())

        # Mese
        tk.Label(
            filter_frame,
            text="Mese:",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).grid(row=0, column=2, sticky="w", padx=(0, 6), pady=10)

        mesi = [label for label, _ in MONTH_CHOICES]
        mese_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.mese_var,
            values=mesi,
            width=14,
            state="readonly",
            font=FONTS["input"],
        )
        mese_combo.grid(row=0, column=3, sticky="ew", padx=(0, 16), pady=10)
        mese_combo.bind("<<ComboboxSelected>>", lambda _: self._carica_premi())

        # Ricerca libera (codice o nome)
        tk.Label(
            filter_frame,
            text="\U0001f50d Ricerca:",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).grid(row=0, column=4, sticky="w", padx=(0, 6), pady=10)

        search_entry = ttk.Entry(
            filter_frame,
            textvariable=self.codice_var,
            width=24,
            font=FONTS["input"],
        )
        search_entry.grid(row=0, column=5, sticky="ew", padx=(0, 16), pady=10)
        search_entry.bind("<Return>", lambda _: self._carica_premi())

        create_button(
            filter_frame,
            text="🔄 Aggiorna",
            command=self._carica_premi,
            variant="secondary",
            width=14,
        ).grid(row=0, column=6, padx=(12, 6), pady=10)

        create_button(
            filter_frame,
            text="🎯 Calcola Premi",
            command=self._genera_premi,
            variant="primary",
            width=16,
        ).grid(row=0, column=7, padx=(6, 12), pady=10)

        create_button(
            filter_frame,
            text="📤 Export Excel",
            command=self._export_premi_excel,
            variant="secondary",
            width=16,
        ).grid(row=0, column=8, padx=(6, 12), pady=10)

        for col in [1, 3, 5]:
            filter_frame.grid_columnconfigure(col, weight=1)

        summary_frame = tk.Frame(filter_frame, bg=COLORS["background"])
        summary_frame.grid(row=1, column=0, columnspan=9, sticky="ew", padx=12, pady=(0, 6))

        tk.Label(
            summary_frame,
            text="Ore Giorno:",
            font=FONTS["label"],
            bg=COLORS["background"],
        ).pack(side="left", padx=(0, 6))

        ore_entry = ttk.Entry(
            summary_frame,
            textvariable=self.ore_giorno_var,
            width=6,
            font=FONTS["input"],
        )
        ore_entry.pack(side="left", padx=(0, 18))

        def _build_summary_block(title: str, var: tk.StringVar) -> None:
            container = tk.Frame(summary_frame, bg=COLORS["background"])
            container.pack(side="left", padx=14)
            tk.Label(
                container,
                text=f"{title}:",
                font=FONTS.get("label", ("Segoe UI", 10, "bold")),
                bg=COLORS["background"],
            ).pack(anchor="w")
            tk.Label(
                container,
                textvariable=var,
                font=FONTS.get("subtitle", ("Segoe UI", 11)),
                bg=COLORS["background"],
                fg=COLORS["primary"],
            ).pack(anchor="w")

        _build_summary_block("Ore MAG", self.total_ore_mag_var)
        _build_summary_block("Ore UFF", self.total_ore_uff_var)
        _build_summary_block("Totale Pallet", self.total_pallet_var)
        _build_summary_block("Media Plt/Ora", self.media_pallet_ora_var)

        # Tabella
        table_frame = tk.Frame(
            self,
            bg=COLORS["white"],
            highlightbackground=COLORS["border"],
            highlightthickness=1,
        )
        table_frame.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        columns = (
            "codice",
            "nome",
            "giorni",
            "giorni_ok",
            "pallet",
            "ore",
            "plt_ora",
            "fascia",
            "premio_base",
            "premio_kpi",
            "premio_totale",
        )

        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
        )

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        headers = {
            "codice": "Codice",
            "nome": "Nome",
            "giorni": "Giorni Lavorati",
            "giorni_ok": "Giorni in Premio",
            "pallet": "Tot. Pallet",
            "ore": "Ore",
            "plt_ora": "Media Plt/h",
            "fascia": "Fascia/Media",
            "premio_base": "Premio Base (€)",
            "premio_kpi": "Premio KPI (€)",
            "premio_totale": "Premio Totale (€)",
        }
        self._tree_headers = headers

        for col, title in headers.items():
            self.tree.heading(col, text=title, command=lambda c=col: self._sort_by_column(c))

        self.tree.column("codice", width=90, anchor="center")
        self.tree.column("nome", width=200, anchor="w")
        self.tree.column("giorni", width=110, anchor="center")
        self.tree.column("giorni_ok", width=130, anchor="center")
        self.tree.column("pallet", width=110, anchor="center")
        self.tree.column("ore", width=90, anchor="center")
        self.tree.column("plt_ora", width=110, anchor="center")
        self.tree.column("fascia", width=140, anchor="center")
        self.tree.column("premio_base", width=120, anchor="e")
        self.tree.column("premio_kpi", width=120, anchor="e")
        self.tree.column("premio_totale", width=140, anchor="e")

        footer_frame = tk.Frame(
            self,
            bg=COLORS["primary"],
        )
        footer_frame.pack(fill="x", padx=16, pady=(4, 0))

        self.stats_label = tk.Label(
            footer_frame,
            text="",
            font=("Segoe UI", 12, "bold"),
            bg=COLORS["primary"],
            fg="white",
        )
        self.stats_label.pack(side="left", padx=12, pady=10)

    # --------------------------------------------------------------- EXPORT ---
    def _collect_tree_rows(self) -> List[Dict[str, str]]:
        columns = list(self.tree["columns"])
        headers: List[str] = []
        for col in columns:
            header = self._tree_headers.get(col)
            headers.append(header if header is not None else col)
        data: List[Dict[str, str]] = []

        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            row = {
                headers[idx]: values[idx] if idx < len(values) else ""
                for idx in range(len(columns))
            }
            data.append(row)

        return data

    # --------------------------------------------------------- ORDINAMENTO ---
    def _sort_by_column(self, col: str) -> None:
        """Ordina la tabella in base alla colonna selezionata."""
        reverse = self._sort_reverse.get(col, False)
        items = list(self.tree.get_children(""))

        def sort_key(item_id: str):
            value = self.tree.set(item_id, col)
            return self._coerce_sort_value(value)

        items.sort(key=sort_key, reverse=reverse)

        for index, item_id in enumerate(items):
            self.tree.move(item_id, "", index)

        self._sort_reverse[col] = not reverse

    @staticmethod
    def _coerce_sort_value(value):
        if value is None:
            return (1, "")
        text = str(value).strip()
        if not text:
            return (1, "")
        normalized = text.replace(" ", "")
        if re.match(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$", normalized):
            normalized = normalized.replace(".", "").replace(",", ".")
        elif re.match(r"^-?\d+([\.,]\d+)?$", normalized):
            normalized = normalized.replace(",", ".")
        try:
            return (0, float(normalized))
        except ValueError:
            return (0, text.lower())

    def _export_premi_excel(self) -> None:
        premi = getattr(self, "_current_premi", [])
        if not premi:
            messagebox.showinfo(
                "Nessun dato",
                "Non ci sono premi da esportare. Usa 'Aggiorna' o 'Calcola Premi'.",
                parent=self,
            )
            return

        anno = self.anno_var.get().strip() or "XXXX"
        mese = (self.mese_var.get().strip() or "Mese").replace(" ", "_")
        default_name = f"Premi_Ricevitori_{anno}_{mese}.xlsx"

        file_path = filedialog.asksaveasfilename(
            title="Esporta premi in Excel",
            defaultextension=".xlsx",
            filetypes=[("File Excel", "*.xlsx"), ("Tutti i file", "*.*")],
            initialdir=str(Path.home()),
            initialfile=default_name,
        )

        if not file_path:
            return

        try:
            import pandas as pd
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter

            def _float_value(value: Any, default: float = 0.0) -> float:
                try:
                    return float(value)
                except (TypeError, ValueError):
                    return default

            data: List[Dict[str, Any]] = []
            for premio in premi:
                media_default = _float_value(premio.get("plt_ora_medio"), 0.0)
                data.append(
                    {
                        "Codice": premio.get("codice_preparatore"),
                        "Nome": premio.get("nome_preparatore"),
                        "Giorni Lavorati": int(premio.get("giorni_lavorati") or 0),
                        "Giorni in Premio": _float_value(premio.get("giorni_in_premio")),
                        "Totale Pallet": _float_value(premio.get("totale_pallet")),
                        "Ore": _float_value(premio.get("ore_lavorate")),
                        "Media Plt/h": media_default,
                        "Fascia/Media": premio.get("fascia_raggiunta")
                        or f"Media {media_default:.2f} Plt/h",
                        "Premio Base (€)": _float_value(premio.get("premio_base")),
                        "Premio KPI (€)": _float_value(premio.get("premio_kpi")),
                        "Premio Totale (€)": _float_value(premio.get("premio_totale")),
                    }
                )

            df = pd.DataFrame(data)

            # Forza tipi numerici e limita a max 2 decimali (evita visualizzazioni tipo ,000)
            int_columns = [
                "Giorni Lavorati",
                "Totale Pallet",
            ]
            dec_columns = [
                "Giorni in Premio",
                "Ore",
                "Media Plt/h",
                "Premio Base (€)",
                "Premio KPI (€)",
                "Premio Totale (€)",
            ]

            for col in int_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).round(0).astype(int)

            for col in dec_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).round(2)

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                sheet_name = "Premi Ricevitori"
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                sheet = writer.sheets[sheet_name]

                header_alignment = Alignment(horizontal="center")
                for cell in sheet[1]:
                    cell.alignment = header_alignment

                # Formati Excel locale-neutral: Excel mostrerà virgola/punto secondo impostazioni
                numeric_formats: Dict[str, str] = {
                    "Giorni Lavorati": "#,##0",
                    "Giorni in Premio": "#,##0.00",
                    "Totale Pallet": "#,##0",
                    "Ore": "#,##0.00",
                    "Media Plt/h": "#,##0.00",
                    "Premio Base (€)": "#,##0.00",
                    "Premio KPI (€)": "#,##0.00",
                    "Premio Totale (€)": "#,##0.00",
                }

                right_alignment = Alignment(horizontal="right")
                left_alignment = Alignment(horizontal="left")
                for col_idx, column_name in enumerate(df.columns, start=1):
                    column_letter = get_column_letter(col_idx)
                    column_values = df[column_name].tolist()
                    text_lengths = [len(str(column_name))]
                    text_lengths.extend(len(str(val)) for val in column_values if val is not None)
                    sheet.column_dimensions[column_letter].width = max(text_lengths or [10]) + 2

                    alignment = right_alignment if column_name in numeric_formats else left_alignment
                    for row_idx in range(2, len(column_values) + 2):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = alignment
                        if column_name in numeric_formats:
                            cell.number_format = numeric_formats[column_name]

                sheet.freeze_panes = "A2"

            open_folder = messagebox.askyesno(
                "Export completato",
                f"File salvato in:\n{file_path}\n\nAprire la cartella di destinazione?",
                parent=self,
            )
            if open_folder:
                try:
                    os.startfile(Path(file_path).parent)
                except OSError:
                    messagebox.showinfo(
                        "Cartella",
                        f"Apri manualmente la cartella:\n{Path(file_path).parent}",
                        parent=self,
                    )
        except Exception as exc:  # pragma: no cover - gestione UI
            messagebox.showerror(
                "Errore",
                f"Esportazione non riuscita:\n{exc}",
                parent=self,
            )

    # --------------------------------------------------------------- DATA ---
    def _carica_premi(self) -> None:
        anno_str = self.anno_var.get().strip()
        mese_label = self.mese_var.get().strip()
        search_text = self.codice_var.get().strip() or None
        self._current_premi = []

        if not anno_str or not mese_label:
            return

        try:
            anno = int(anno_str)
            mese = next(m for label, m in MONTH_CHOICES if label == mese_label)
        except (ValueError, StopIteration):
            return

        PremiRicevitoriView._saved_state = {
            "search": self.codice_var.get().strip()
        }

        try:
            premi = fetch_premi_ricevitori(anno, mese, search=search_text)
            self._current_premi = premi.copy()

            for item in self.tree.get_children():
                self.tree.delete(item)

            totale_premi = Decimal("0")
            totale_giorni_ok = Decimal("0")
            bonus_applicato = False

            for premio in premi:
                giorni = int(premio.get("giorni_lavorati") or 0)
                giorni_ok = Decimal(str(premio.get("giorni_in_premio") or 0))
                pallet = Decimal(str(premio.get("totale_pallet") or 0))
                ore = Decimal(str(premio.get("ore_lavorate") or 0))
                plt_ora = Decimal(str(premio.get("plt_ora_medio") or 0))
                premio_base = Decimal(str(premio.get("premio_base") or 0))
                premio_kpi = Decimal(str(premio.get("premio_kpi") or 0))
                premio_totale = Decimal(str(premio.get("premio_totale") or 0))

                if premio.get("bonus_applicato") or premio_kpi > 0:
                    bonus_applicato = True

                self.tree.insert(
                    "",
                    "end",
                    values=(
                        premio.get("codice_preparatore"),
                        premio.get("nome_preparatore"),
                        giorni,
                        f"{giorni_ok:.2f}",
                        f"{pallet:,.0f}",
                        f"{ore:.2f}",
                        f"{plt_ora:.2f}",
                        premio.get("fascia_raggiunta") or "-",
                        f"{premio_base:.2f}",
                        f"{premio_kpi:.2f}",
                        f"{premio_totale:.2f}",
                    ),
                )

                totale_premi += premio_totale
                totale_giorni_ok += giorni_ok

            codici_presenti: List[str] = sorted(
                {
                    str(premio.get("codice_preparatore") or "").upper()
                    for premio in premi
                    if premio.get("codice_preparatore")
                }
            )
            if not codici_presenti:
                codici_presenti = self._collect_codici_ricevitori(anno, mese, search_text)

            totale_pallet_lordo = self._fetch_totale_pallet_lordo(anno, mese, search_text)
            ore_mag, ore_uff = self._fetch_totali_ore_mag_uff(anno, mese, codici_presenti)

            if not premi:
                self.stats_label.config(
                    text=f"Nessun premio calcolato per {mese_label} {anno}. Usa 'Calcola Premi' per generarli."
                )
                self._update_totals_summary(ore_mag, ore_uff, totale_pallet_lordo)
            else:
                bonus_text = "✓" if bonus_applicato else "✗"
                self.stats_label.config(
                    text=(
                        f"Ricevitori: {len(premi)} | Giorni in premio: {totale_giorni_ok:.2f} "
                        f"| Totale Premi: €{totale_premi:,.2f} | Bonus KPI 15%: {bonus_text}"
                    )
                )
                self._update_totals_summary(ore_mag, ore_uff, totale_pallet_lordo)
        except Exception as exc:  # pragma: no cover - gestione UI
            messagebox.showerror(
                "Errore",
                f"Errore nel caricamento premi:\n{exc}",
                parent=self,
            )
            self._update_totals_summary(Decimal("0"), Decimal("0"), Decimal("0"))
            self._current_premi = []

    def _update_totals_summary(
        self,
        ore_mag: Decimal,
        ore_uff: Decimal,
        totale_pallet: Decimal,
    ) -> None:
        try:
            ore_mag_val = max(float(ore_mag), 0.0)
        except (TypeError, ValueError, ArithmeticError):
            ore_mag_val = 0.0

        try:
            ore_uff_val = max(float(ore_uff), 0.0)
        except (TypeError, ValueError, ArithmeticError):
            ore_uff_val = 0.0

        try:
            pallet_val = max(float(totale_pallet), 0.0)
        except (TypeError, ValueError, ArithmeticError):
            pallet_val = 0.0

        media = (pallet_val / ore_mag_val) if ore_mag_val > 0 else 0.0

        self.total_ore_mag_var.set(f"{ore_mag_val:.2f}")
        self.total_ore_uff_var.set(f"{ore_uff_val:.2f}")
        self.total_pallet_var.set(f"{pallet_val:,.0f}")
        self.media_pallet_ora_var.set(f"{media:.2f}")

    def _fetch_totale_pallet_lordo(
        self,
        anno: int,
        mese: int,
        codice: Optional[str],
    ) -> Decimal:
        tipo_validi = ("RICEVITORI", *LOCAL_TIPI_UFF)
        placeholders = ", ".join(["%s"] * len(tipo_validi))
        query = f"""
            SELECT COALESCE(SUM(dp.totale_colli), 0)
            FROM dati_produzione dp
            WHERE dp.tipo_attivita IN ({placeholders})
              AND YEAR(dp.data) = %s
              AND MONTH(dp.data) = %s
              AND NOT EXISTS (
                  SELECT 1 FROM anomalie a
                  WHERE a.data_rilevamento = dp.data
                      AND a.codice_preparatore = dp.codice_preparatore
                      AND a.tipo_attivita = dp.tipo_attivita
                      AND a.stato = 'ELIMINATA'
              )
        """
        params: List[Any] = [*tipo_validi, anno, mese]

        codice_norm = (codice or "").strip().upper()
        if codice_norm:
            query += " AND UPPER(dp.codice_preparatore) = %s"
            params.append(codice_norm)

        row = None
        try:
            with closing(mysql.connector.connect(**MYSQL_CONFIG)) as conn:
                with closing(conn.cursor()) as cur:
                    cur.execute(query, params)
                    row = cur.fetchone()
        except mysql.connector.Error as exc:
            print(f"[PremiRicevitori] Errore nel calcolo totale pallet lordo: {exc}")
            return Decimal("0")

        totale = row[0] if row else 0  # type: ignore[index]
        try:
            return Decimal(str(totale or 0))
        except Exception:
            return Decimal("0")

    def _fetch_totali_ore_mag_uff(
        self,
        anno: int,
        mese: int,
        codici: List[str],
    ) -> Tuple[Decimal, Decimal]:
        if not codici:
            return Decimal("0"), Decimal("0")

        tipo_mag = TIM_TIPI_MAG
        tipo_uff = TIM_TIPI_UFF
        tipo_all = tipo_mag + tipo_uff

        first_day = datetime.date(anno, mese, 1)
        last_day = datetime.date(anno, mese, calendar.monthrange(anno, mese)[1])

        code_placeholders = ", ".join(["%s"] * len(codici))
        tipo_placeholders = ", ".join(["%s"] * len(tipo_all))

        query = f"""
            SELECT LOWER(cg.codice) AS codice,
                   ta.descrizione AS tipo,
                   COALESCE(
                       SUM(
                           CASE
                               WHEN a.data_inizio IS NOT NULL AND a.data_fine IS NOT NULL THEN GREATEST(
                                   TIMESTAMPDIFF(MINUTE, a.data_inizio, a.data_fine)
                                     - COALESCE((SELECT SUM(p.durata) FROM pausa p WHERE p.attivita_id = a.id), 0),
                                   0
                               )
                               ELSE 0
                           END
                       ),
                       0
                   ) AS durata_totale
            FROM attivita a
            JOIN utente u ON a.utente_id = u.id
            JOIN codicegestionale cg ON cg.utente_id = u.id
                AND LOWER(cg.codice) IN ({code_placeholders})
                AND a.data_riferimento BETWEEN cg.valido_dal AND COALESCE(cg.valido_al, '9999-12-31')
            JOIN tipoattivita ta ON a.tipo_attivita_id = ta.id
            WHERE a.data_riferimento BETWEEN %s AND %s
              AND ta.descrizione IN ({tipo_placeholders})
            GROUP BY codice, tipo
        """

        params: List[Any] = [code.lower() for code in codici]
        params.extend([first_day, last_day])
        params.extend(tipo_all)

        ore_mag = Decimal("0")
        ore_uff = Decimal("0")
        tipo_mag_upper = {t.upper() for t in tipo_mag}
        tipo_uff_upper = {t.upper() for t in tipo_uff}

        try:
            with closing(mysql.connector.connect(**MYSQL_CONFIG_MAIN)) as conn:
                with closing(conn.cursor(dictionary=True)) as cur:
                    cur.execute(query, params)
                    rows = cur.fetchall()
        except mysql.connector.Error as exc:
            print(f"[PremiRicevitori] Errore nel calcolo ore MAG/UFF: {exc}")
            return Decimal("0"), Decimal("0")

        for raw_row in rows or []:
            row = cast(Dict[str, Any], raw_row)
            tipo = (row.get("tipo") or "").upper()
            minuti_tot = Decimal(str(row.get("durata_totale") or 0))
            ore = minuti_tot / Decimal("60")
            if tipo in tipo_mag_upper:
                ore_mag += ore
            elif tipo in tipo_uff_upper:
                ore_uff += ore

        return ore_mag, ore_uff

    def _collect_codici_ricevitori(
        self,
        anno: int,
        mese: int,
        codice: Optional[str],
    ) -> List[str]:
        tipo_validi = ("RICEVITORI", *LOCAL_TIPI_UFF)
        placeholders = ", ".join(["%s"] * len(tipo_validi))
        query = f"""
            SELECT DISTINCT UPPER(dp.codice_preparatore) AS codice
            FROM dati_produzione dp
            WHERE dp.tipo_attivita IN ({placeholders})
              AND YEAR(dp.data) = %s
              AND MONTH(dp.data) = %s
        """
        params: List[Any] = [*tipo_validi, anno, mese]

        codice_norm = (codice or "").strip().upper()
        if codice_norm:
            query += " AND UPPER(dp.codice_preparatore) = %s"
            params.append(codice_norm)

        try:
            with closing(mysql.connector.connect(**MYSQL_CONFIG)) as conn:
                with closing(conn.cursor(dictionary=True)) as cur:
                    cur.execute(query, params)
                    rows = cur.fetchall()
        except mysql.connector.Error as exc:
            print(f"[PremiRicevitori] Errore nel recupero codici ricevimento: {exc}")
            return []

        codici: List[str] = []
        for raw_row in rows or []:
            row = cast(Dict[str, Any], raw_row)
            codice_val = str(row.get("codice") or "").upper()
            if codice_val:
                codici.append(codice_val)
        return codici

    def _genera_premi(self) -> None:
        anno_str = self.anno_var.get().strip()
        mese_label = self.mese_var.get().strip()

        if not anno_str or not mese_label:
            messagebox.showwarning(
                "Dati mancanti",
                "Seleziona anno e mese.",
                parent=self,
            )
            return

        try:
            anno = int(anno_str)
            mese = next(m for label, m in MONTH_CHOICES if label == mese_label)
        except (ValueError, StopIteration):
            messagebox.showerror("Errore", "Anno o mese non validi.", parent=self)
            return

        premi_esistenti = fetch_premi_ricevitori(anno, mese)
        if premi_esistenti:
            risposta = messagebox.askyesno(
                "Conferma",
                (
                    f"Esistono già {len(premi_esistenti)} premi calcolati per {mese_label} {anno}.\n\n"
                    "Vuoi ricalcolarli? I dati precedenti verranno sovrascritti."
                ),
                parent=self,
            )
            if not risposta:
                return

        try:
            fasce = self._load_fasce_premio()
            if not fasce:
                messagebox.showwarning(
                    "Fasce mancanti",
                    "Non sono definite fasce premio per RICEVITORI.",
                    parent=self,
                )
                return

            bonus_perc = self._load_bonus_malus(anno, mese)

            risultati = self._calcola_premi_ricevitori(
                anno=anno,
                mese=mese,
                codice_filtro=None,
                fasce=fasce,
                bonus_perc=bonus_perc,
            )

            if not risultati:
                messagebox.showinfo(
                    "Nessun dato",
                    f"Nessun dato di produzione trovato per {mese_label} {anno}.",
                    parent=self,
                )
                return

            save_premi_ricevitori(anno, mese, risultati)

            messagebox.showinfo(
                "Successo",
                f"Calcolati e salvati {len(risultati)} premi per {mese_label} {anno}.",
                parent=self,
            )

            self._carica_premi()
        except Exception as exc:  # pragma: no cover - gestione UI
            messagebox.showerror(
                "Errore",
                f"Errore nel calcolo premi:\n{exc}",
                parent=self,
            )

    # ----------------------------------------------------------- DATA HELP ---
    def _load_fasce_premio(self) -> List[Dict[str, Any]]:
        fasce = fetch_fasce_premi("RICEVITORI")
        return sorted(fasce, key=lambda f: f.get("valore_riferimento", 0))

    def _load_bonus_malus(self, anno: int, mese: int) -> Optional[Decimal]:
        record = get_malus_bonus(anno, mese)
        if not record:
            return None

        attivita_bonus = (record.get("attivita_bonus") or "").upper()
        if "RICEVITORI" not in attivita_bonus:
            return None

        try:
            rotture = Decimal(str(record.get("importo_rotture", 0) or 0))
            differenze = Decimal(str(record.get("importo_differenze", 0) or 0))
            soglia_rot = Decimal(str(record.get("soglia_rotture", 0) or 0))
            soglia_diff = Decimal(str(record.get("soglia_differenze", 0) or 0))

            # Il bonus scatta solo se ENTRAMBE le soglie sono rispettate
            # (verifica separata, non aggregata)
            if rotture <= soglia_rot and differenze <= soglia_diff:
                return Decimal("0.15")
        except Exception:
            return None

        return None

    def _calcola_premi_ricevitori(
        self,
        anno: int,
        mese: int,
        codice_filtro: Optional[str],
        fasce: List[Dict[str, Any]],
        bonus_perc: Optional[Decimal],
    ) -> List[Dict[str, Any]]:
        if DEBUG_PREPARATORI_ORE:
            print(
                "[PremiRicevitori] Debug ore attivo per: "
                + ", ".join(sorted(DEBUG_PREPARATORI_ORE))
            )
        cent = Decimal("0.01")
        fasce_sorted = [
            {
                "soglia": Decimal(str(f.get("valore_riferimento", 0))),
                "premio": Decimal(str(f.get("valore_premio", 0))),
                "unita": str(f.get("unita_riferimento") or "Plt/h"),
            }
            for f in fasce
        ]
        fasce_sorted.sort(key=lambda item: item["soglia"])

        tipo_validi = ("RICEVITORI", *LOCAL_TIPI_UFF)
        placeholders = ", ".join(["%s"] * len(tipo_validi))

        query = f"""
            SELECT
                dp.data,
                dp.codice_preparatore,
                dp.nome_preparatore,
                SUM(dp.totale_colli) AS totale_pallet
            FROM dati_produzione dp
            WHERE dp.tipo_attivita IN ({placeholders})
                AND YEAR(dp.data) = %s
                AND MONTH(dp.data) = %s
                AND NOT EXISTS (
                    SELECT 1 FROM anomalie a
                    WHERE a.data_rilevamento = dp.data
                        AND a.codice_preparatore = dp.codice_preparatore
                        AND a.tipo_attivita = dp.tipo_attivita
                        AND a.stato = 'ELIMINATA'
                )
        """
        params: List[Any] = [*tipo_validi, anno, mese]

        if codice_filtro:
            query += " AND dp.codice_preparatore = %s"
            params.append(codice_filtro)

        query += " GROUP BY dp.data, dp.codice_preparatore, dp.nome_preparatore"

        risultati: Dict[str, Dict[str, Any]] = {}
        local_rows: List[Dict[str, Any]] = []

        with closing(mysql.connector.connect(**MYSQL_CONFIG)) as conn:
            with closing(conn.cursor(dictionary=True)) as cur:
                cur.execute(query, params)
                rows = cur.fetchall()

                for raw_row in rows:
                    row = cast(Dict[str, Any], raw_row)
                    codice = str(row.get("codice_preparatore") or "").strip()
                    if not codice:
                        continue
                    nome = str(row.get("nome_preparatore") or "").strip()
                    data_val = row.get("data")
                    if isinstance(data_val, datetime.datetime):
                        data_obj = data_val.date()
                    elif isinstance(data_val, datetime.date):
                        data_obj = data_val
                    else:
                        try:
                            data_obj = datetime.datetime.strptime(str(data_val), "%Y-%m-%d").date()
                        except Exception:
                            continue

                    pallet = Decimal(str(row.get("totale_pallet") or 0))
                    if pallet < 0:
                        continue

                    local_rows.append(
                        {
                            "codice": codice,
                            "nome": nome,
                            "data": data_obj,
                            "pallet": pallet,
                        }
                    )

        if not local_rows:
            return []

        days_in_month = calendar.monthrange(anno, mese)[1]
        month_dates = [
            datetime.date(anno, mese, day).strftime("%Y-%m-%d") for day in range(1, days_in_month + 1)
        ]

        durations_map, tim_name_map, uff_map = self._fetch_tim_minutes(
            local_rows, MYSQL_CONFIG_MAIN, month_dates
        )

        # Premio di squadra: determina la fascia dal valore complessivo del mese
        # (media Plt/Ora = Totale pallet / Ore MAG), coerente con il riepilogo in UI.
        totale_pallet_squadra = sum((row.get("pallet") or Decimal("0")) for row in local_rows)
        minuti_totali = sum(durations_map.values(), Decimal("0"))
        minuti_uff = sum(uff_map.values(), Decimal("0"))
        minuti_mag = minuti_totali - minuti_uff
        ore_mag_squadra = (minuti_mag / Decimal("60")) if minuti_mag > 0 else Decimal("0")
        media_squadra = (
            (totale_pallet_squadra / ore_mag_squadra).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if ore_mag_squadra > 0
            else Decimal("0")
        )

        fascia_squadra = None
        for fascia_def in fasce_sorted:
            if media_squadra >= fascia_def["soglia"]:
                fascia_squadra = fascia_def
        premio_giorno_squadra = (fascia_squadra["premio"] if fascia_squadra else Decimal("0")).quantize(cent)
        fascia_label_squadra = f"Media squadra {media_squadra:.2f} Plt/h"

        existing_keys = {
            (row["codice"].upper(), row["data"].strftime("%Y-%m-%d")) for row in local_rows
        }

        for (code, data_str), minuti_totali in durations_map.items():
            if (code, data_str) in existing_keys or minuti_totali <= 0:
                continue
            try:
                data_obj = datetime.datetime.strptime(data_str, "%Y-%m-%d").date()
            except ValueError:
                continue
            local_rows.append(
                {
                    "codice": code,
                    "nome": tim_name_map.get((code, data_str), ""),
                    "data": data_obj,
                    "pallet": Decimal("0"),
                }
            )
            existing_keys.add((code, data_str))

        debug_entries: List[Dict[str, Any]] = []

        for row in local_rows:
            codice = row["codice"].upper()
            nome = row["nome"]
            data_str = row["data"].strftime("%Y-%m-%d")
            minuti = durations_map.get((codice, data_str), Decimal("0"))
            if minuti <= 0:
                continue

            ore = minuti / Decimal("60")
            if ore <= 0:
                continue

            pallet = row["pallet"]

            # Manteniamo la media individuale per statistica/colonna, ma il premio è di squadra.
            plt_ora = (pallet / ore).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            premio_giorno = premio_giorno_squadra

            display_nome = nome or tim_name_map.get((codice, data_str), "")

            stats = risultati.setdefault(
                codice,
                {
                    "codice": codice,
                    "nome": display_nome,
                    "giorni_lavorati": 0,
                    "giorni_in_premio": 0,
                    "tot_pallet": Decimal("0"),
                    "ore": Decimal("0"),
                    "premio_base": Decimal("0"),
                    "fascia_max": Decimal("0"),
                    "fascia_label": fascia_label_squadra,
                },
            )

            stats["nome"] = stats["nome"] or display_nome
            stats["giorni_lavorati"] += 1
            stats["tot_pallet"] += pallet
            stats["ore"] += ore

            if self._is_debug_preparatore(codice, display_nome):
                debug_entries.append(
                    {
                        "data": data_str,
                        "nome": display_nome or codice,
                        "minuti": minuti,
                        "ore": ore,
                        "pallet": pallet,
                        "media": plt_ora,
                    }
                )
            # Nota: i giorni in premio vengono calcolati a fine mese come ore_totali / ore_giorno.

        if debug_entries:
            debug_entries.sort(key=lambda entry: entry["data"])
            print("[PremiRicevitori] Dettaglio giornaliero (ordinato):")
            for entry in debug_entries:
                print(
                    "[PremiRicevitori][DEBUG ORE] "
                    f"{entry['nome']} | Giorno {entry['data']} | Minuti TIM: {float(entry['minuti']):.2f} "
                    f"(Ore {float(entry['ore']):.2f}) | Pallet: {float(entry['pallet']):.2f} "
                    f"| Media: {float(entry['media']):.2f}"
                )

        risultati_finali: List[Dict[str, Any]] = []
        for stats in risultati.values():
            ore = stats["ore"]
            if ore <= 0:
                continue

            # Giorni in premio: calcolati dalle ore totali e dal parametro "Ore Giorno"
            try:
                ore_giorno = Decimal(str(self.ore_giorno_var.get() or "8")).quantize(Decimal("0.01"))
            except Exception:
                ore_giorno = Decimal("8")
            if ore_giorno <= 0:
                ore_giorno = Decimal("8")

            giorni_in_premio = (ore / ore_giorno).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if giorni_in_premio < 0:
                giorni_in_premio = Decimal("0")

            plt_medio = (stats["tot_pallet"] / ore).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            premio_base = (premio_giorno_squadra * giorni_in_premio).quantize(cent)

            premio_kpi = Decimal("0")
            if bonus_perc and premio_base > 0:
                premio_kpi = (premio_base * bonus_perc).quantize(cent)

            premio_totale = (premio_base + premio_kpi).quantize(cent)

            if self._is_debug_preparatore(stats["codice"], stats["nome"]):
                print(
                    "[PremiRicevitori][DEBUG ORE - TOTALE] "
                    f"{stats['nome'] or stats['codice']} | Giorni: {stats['giorni_lavorati']} "
                    f"(In premio: {giorni_in_premio:.2f}) | Ore: {float(ore):.2f} "
                    f"| Pallet: {float(stats['tot_pallet']):.2f} | Media: {float(plt_medio):.2f} "
                    f"| Premio base: {float(premio_base):.2f} | Fascia: {stats['fascia_label'] or 'N/D'}"
                )

            risultati_finali.append(
                {
                    "codice": stats["codice"],
                    "nome": stats["nome"],
                    "giorni_lavorati": stats["giorni_lavorati"],
                    "giorni_in_premio": giorni_in_premio,
                    "tot_pallet": float(stats["tot_pallet"]),
                    "ore": float(ore),
                    "plt_ora": float(plt_medio),
                    "fascia": stats["fascia_label"] or f"Media {plt_medio:.2f} Plt/h",
                    "premio_base": float(premio_base),
                    "premio_kpi": float(premio_kpi),
                    "premio_totale": float(premio_totale),
                    "bonus_applicato": bool(premio_kpi > 0),
                    "note": None,
                }
            )

        risultati_finali.sort(key=lambda item: item["premio_totale"], reverse=True)
        return risultati_finali

    def _fetch_tim_minutes(
        self,
        local_rows: List[Dict[str, Any]],
        mysql_config_main: Dict[str, Any],
        date_list: List[str],
    ) -> Tuple[
        Dict[tuple[str, str], Decimal],
        Dict[tuple[str, str], str],
        Dict[tuple[str, str], Decimal],
    ]:
        if not local_rows or not date_list:
            return {}, {}, {}

        tipo_tim = TIM_TIPI_MAG + TIM_TIPI_UFF
        tipo_uff_upper = {t.upper() for t in TIM_TIPI_UFF}
        tipo_placeholders = ", ".join(["%s"] * len(tipo_tim))
        tipo_codice_placeholders = ", ".join(["%s"] * len(tipo_tim))

        durate_map: Dict[tuple[str, str], Decimal] = {}
        nome_map: Dict[tuple[str, str], str] = {}
        uff_map: Dict[tuple[str, str], Decimal] = {}

        query = f"""
            SELECT LOWER(cg.codice) AS codice,
                   ta.descrizione AS tipo,
                   u.nome,
                   u.cognome,
                   COALESCE(
                       SUM(
                           CASE
                               WHEN a.data_inizio IS NOT NULL AND a.data_fine IS NOT NULL THEN GREATEST(
                                   TIMESTAMPDIFF(MINUTE, a.data_inizio, a.data_fine)
                                     - COALESCE((SELECT SUM(p.durata) FROM pausa p WHERE p.attivita_id = a.id), 0),
                                   0
                               )
                               ELSE 0
                           END
                       ),
                       0
                   ) AS durata_totale
            FROM attivita a
            JOIN utente u ON a.utente_id = u.id
            JOIN tipoattivita ta ON a.tipo_attivita_id = ta.id
            JOIN codicegestionale cg ON cg.utente_id = u.id
                AND a.data_riferimento BETWEEN cg.valido_dal AND COALESCE(cg.valido_al, '9999-12-31')
            JOIN tipoattivita ta_codice ON ta_codice.id = cg.tipo_attivita_id
                AND ta_codice.descrizione IN ({tipo_codice_placeholders})
            WHERE a.data_riferimento = %s
              AND ta.descrizione IN ({tipo_placeholders})
            GROUP BY codice, tipo, u.nome, u.cognome
        """

        with closing(mysql.connector.connect(**mysql_config_main)) as tim_conn:
            with closing(tim_conn.cursor(dictionary=True)) as tim_cursor:
                for data_str in date_list:
                    params: List[Any] = list(tipo_tim)
                    params.append(data_str)
                    params.extend(tipo_tim)

                    tim_cursor.execute(query, params)
                    rows = tim_cursor.fetchall()

                    for raw_row in rows:
                        row = cast(Dict[str, Any], raw_row)
                        code = str(row.get("codice") or "").upper()
                        durata = Decimal(str(row.get("durata_totale") or 0))
                        key = (code, data_str)
                        durate_map[key] = durate_map.get(key, Decimal("0")) + durata

                        tipo_descr = str(row.get("tipo") or "").upper()
                        if tipo_descr in tipo_uff_upper and durata > 0:
                            uff_map[key] = uff_map.get(key, Decimal("0")) + durata

                        nome = str(row.get("nome") or "").strip()
                        cognome = str(row.get("cognome") or "").strip()
                        nominativo = " ".join(part for part in [cognome.upper(), nome.upper()] if part)
                        if nominativo:
                            nome_map.setdefault(key, nominativo.strip())

        return durate_map, nome_map, uff_map

    def _is_debug_preparatore(self, codice: Optional[str], nome: Optional[str]) -> bool:
        target = DEBUG_PREPARATORI_ORE
        if not target:
            return False
        nome_norm = (nome or "").strip().upper()
        codice_norm = (codice or "").strip().upper()
        for raw in target:
            needle = raw.strip().upper()
            if not needle:
                continue
            if nome_norm and needle in nome_norm:
                return True
            if codice_norm and codice_norm == needle:
                return True
        return False


__all__ = ["PremiRicevitoriView"]
